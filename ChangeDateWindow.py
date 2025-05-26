import os
from PyQt5.QtWidgets import (QLabel, QVBoxLayout, QDialog, QSpinBox, QPushButton, 
                            QMessageBox, QCalendarWidget, QHBoxLayout, QComboBox, 
                            QFrame, QGraphicsDropShadowEffect)
from PyQt5.QtGui import QIcon, QColor, QFont, QPalette
from PyQt5.QtCore import Qt, QDateTime, pyqtSignal, QLocale
from openpyxl import load_workbook
from datetime import datetime
from copticDate import CopticCalendar
from commonFunctions import relative_path

class ChangeDate(QDialog):
    date_updated = pyqtSignal(str, str)

    def __init__(self, CurrentDate, CurrentTime):
        super().__init__()

        try:
            # Set window properties
            self.setWindowTitle("إختيار التاريخ")
            self.setWindowIcon(QIcon(relative_path(r"Data\الصور\Logo.ico")))
            self.setStyleSheet("""
                QDialog {
                    background-color: #f0f5ff;
                }
                QLabel {
                    color: #1a365d;
                    font-weight: bold;
                    font-size: 14px;
                    margin-top: 5px;
                }
                QPushButton {
                    background-color: #1a365d;
                    color: white;
                    border-radius: 5px;
                    padding: 8px;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton:hover {
                    background-color: #2a466d;
                }
                QPushButton:pressed {
                    background-color: #0a264d;
                }
                QComboBox, QSpinBox {
                    border: 1px solid #1a365d;
                    border-radius: 5px;
                    padding: 5px;
                    background-color: white;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 24px;
                }
                QSpinBox::up-button, QSpinBox::down-button {
                    width: 20px;
                    color: white;
                }
            """)

            # Load options from Excel and set up ComboBox
            self.load_options_from_excel()
            self.excel_dropdown = QComboBox()
            self.excel_dropdown.setStyleSheet("""
                QComboBox {
                    font-size: 14px;
                    min-height: 30px;
                }
            """)
            
            # Add a placeholder value to the dropdown menu
            self.excel_dropdown.addItem("المناسبات")
            self.excel_dropdown.addItems(self.options)
            self.excel_dropdown.setCurrentIndex(0)

            # Create styled buttons
            self.submit_button = QPushButton("تعين")
            self.submit_button.clicked.connect(self.submit_date_time)
            
            self.current_date_time_button = QPushButton("مبـــاشر")
            self.current_date_time_button.clicked.connect(self.show_current_date_time)
            
            # Apply button effects
            for button in [self.submit_button, self.current_date_time_button]:
                shadow = QGraphicsDropShadowEffect()
                shadow.setBlurRadius(10)
                shadow.setColor(QColor(0, 0, 0, 60))
                shadow.setOffset(2, 2)
                button.setGraphicsEffect(shadow)

            # Create styled Coptic date label
            self.coptic_date_label = QLabel()
            self.coptic_date_label.setStyleSheet("""
                QLabel {
                    background-color: #e6eeff;
                    border: 1px solid #1a365d;
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 16px;
                    font-weight: bold;
                    color: #1a365d;
                }
            """)

            # Create CopticCalendar instance
            self.coptic_calendar = CopticCalendar()

            # Create the main layout
            layout = QVBoxLayout()
            layout.setSpacing(10)
            layout.setContentsMargins(15, 15, 15, 15)

            # Create header
            header_label = QLabel("اختر التاريخ:")
            header_label.setStyleSheet("font-size: 16px; color: #1a365d; font-weight: bold;")
            layout.addWidget(header_label)

            # Create calendar container frame
            calendar_frame = QFrame()
            calendar_frame.setStyleSheet("""
                QFrame {
                    border: 1px solid #1a365d;
                    border-radius: 8px;
                    background-color: white;
                }
            """)
            calendar_layout = QVBoxLayout(calendar_frame)
            calendar_layout.setContentsMargins(0, 0, 0, 0)  # Remove inner padding
            
            # Create and style the calendar widget
            self.calendar_widget = QCalendarWidget()
            self.calendar_widget.setGridVisible(True)
            self.calendar_widget.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
            self.calendar_widget.setLocale(QLocale(QLocale.Arabic, QLocale.Egypt))
            self.calendar_widget.setSelectedDate(CurrentDate)
            self.calendar_widget.clicked.connect(self.update_coptic_date_label)
            
            # Style the calendar - remove the inner border by setting it to transparent
            self.calendar_widget.setStyleSheet("""
                QCalendarWidget {
                    background-color: transparent;
                    color: #1a365d;
                    border: none;
                }
                QCalendarWidget QToolButton {
                    color: #1a365d;
                    background-color: #e6eeff;
                    font-size: 14px;
                }
                QCalendarWidget QMenu {
                    color: #1a365d;
                    background-color: white;
                }
                QCalendarWidget QSpinBox {
                    color: #1a365d;
                    background-color: white;
                }
                QCalendarWidget QWidget#qt_calendar_navigationbar {
                    background-color: #e6eeff;
                    border-top-left-radius: 8px;
                    border-top-right-radius: 8px;
                }
                QCalendarWidget QWidget { 
                    alternate-background-color: #f0f5ff; 
                }
                QCalendarWidget QAbstractItemView:enabled {
                    color: #1a365d;
                    selection-background-color: #1a365d;
                    selection-color: white;
                    border: none;
                }
                QCalendarWidget QAbstractItemView:disabled {
                    color: gray;
                }
            """)

            calendar_layout.addWidget(self.calendar_widget)
            layout.addWidget(calendar_frame)

            # Time section
            time_label = QLabel("اختر الوقت (اليوم القبطي ينتهي 5:30 مساء)")
            time_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(time_label)

            # Create styled time selection container
            time_frame = QFrame()
            time_frame.setStyleSheet("""
                QFrame {
                    border: 1px solid #1a365d;
                    border-radius: 5px;
                    background-color: white;
                    padding: 10px;
                }
            """)
            time_layout = QHBoxLayout(time_frame)
            time_layout.setSpacing(10)  # Add spacing between elements

            # Style the spinboxes with modern up/down buttons
            spinbox_style = """
                QSpinBox {
                    border: 1px solid #1a365d;
                    border-radius: 5px;
                    padding: 5px;
                    font-size: 16px;
                    min-width: 60px;
                    min-height: 30px;
                    background-color: white;
                    selection-background-color: #1a365d;
                    selection-color: white;
                }
            """

            # Create a custom spinbox class for minutes to show "00" instead of "0"
            class MinuteSpinBox(QSpinBox):
                def textFromValue(self, value):
                    return f"{value:02d}"  # Format with leading zero

            # Create SpinBoxes for hours and minutes
            self.hour_spin = QSpinBox()
            self.hour_spin.setRange(0, 13)
            self.hour_spin.setValue(int(CurrentTime.split(':')[0]) % 12 or 12)
            self.hour_spin.valueChanged.connect(self.handle_hour_change)
            self.hour_spin.setStyleSheet(spinbox_style)
            self.hour_spin.setAlignment(Qt.AlignCenter)  # Center align text

            # Add colon label between hour and minute - removed borders/background
            colon_label = QLabel(":")
            colon_label.setStyleSheet("""
                QLabel {
                    font-size: 24px;
                    font-weight: bold;
                    color: #1a365d;
                    background: transparent;
                    border: none;
                    margin: 0 2px;
                    padding: 0;
                }
            """)
            colon_label.setAlignment(Qt.AlignCenter)
            
            # Use custom MinuteSpinBox instead of regular QSpinBox
            self.minute_spin = MinuteSpinBox()
            self.minute_spin.setRange(0, 60)
            self.minute_spin.setValue(int(CurrentTime.split(':')[1].split()[0]))
            self.minute_spin.valueChanged.connect(self.handle_minute_change)
            self.minute_spin.setStyleSheet(spinbox_style)
            self.minute_spin.setAlignment(Qt.AlignCenter)  # Center align text

            # AM/PM Combo
            self.ampm_combo = QComboBox()
            self.ampm_combo.addItems(["AM", "PM"])
            self.ampm_combo.setCurrentIndex(0 if "AM" in CurrentTime else 1)
            self.ampm_combo.setStyleSheet("""
                QComboBox {
                    border: 1px solid #1a365d;
                    border-radius: 5px;
                    padding: 5px;
                    font-size: 16px;
                    min-width: 70px;
                    min-height: 30px;
                    text-align: center;
                }
                QComboBox::drop-down {
                    subcontrol-origin: padding;
                    subcontrol-position: center right;
                    width: 24px;
                    border-left-width: 1px;
                    border-left-color: #1a365d;
                    border-left-style: solid;
                    border-top-right-radius: 5px;
                    border-bottom-right-radius: 5px;
                    color: #1a365d;
                }
            """)
            self.ampm_combo.setEditable(False)

            self.ampm_combo.currentIndexChanged.connect(self.update_coptic_date_label)  # Default to AM

            # Add widgets to time layout
            time_layout.addStretch(1)  # Add stretch to center the time controls
            time_layout.addWidget(self.hour_spin)
            time_layout.addWidget(colon_label)
            time_layout.addWidget(self.minute_spin)
            time_layout.addWidget(self.ampm_combo)
            time_layout.addStretch(1)  # Add stretch to center the time controls            
            layout.addWidget(time_frame)

            # Occasion selection section
            occasion_label = QLabel("اختار المنــاسبة")
            occasion_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(occasion_label)
            layout.addWidget(self.excel_dropdown)
            
            # Add Coptic date display
            layout.addWidget(self.coptic_date_label)
            
            # Button layout
            button_layout = QHBoxLayout()
            button_layout.addWidget(self.current_date_time_button)
            button_layout.addWidget(self.submit_button)
            layout.addLayout(button_layout)

            self.setLayout(layout)

            # Connect ComboBox signal
            self.excel_dropdown.currentIndexChanged.connect(self.update_date_and_time_from_coptic)
            self.update_coptic_date_label()
        except Exception as e:
            self.show_error_message(str(e))

    def load_options_from_excel(self):
        try:
            # Load options from Excel
            self.options = []
            workbook_path = relative_path(r"Tables.xlsx")
            wb = load_workbook(workbook_path)
            ws = wb["المناسبات"]
            for cell in ws['G'][1:]:
                if cell.value:
                    self.options.append(cell.value)
        except Exception as e:
            self.show_error_message(str(e))

    def show_current_date_time(self):
        try:
            # Get current date and time
            current_date_time = QDateTime.currentDateTime()
            formatted_date_time = current_date_time.toString(Qt.ISODate)
            selected_date = QDateTime.fromString(formatted_date_time.split("T")[0], Qt.ISODate).date()
            self.calendar_widget.setSelectedDate(selected_date)  # Set the current date in the calendar
            self.hour_spin.setValue(current_date_time.time().hour() % 12 or 12)  # Convert to 12-hour format
            self.minute_spin.setValue(current_date_time.time().minute())
            self.ampm_combo.setCurrentIndex(0 if current_date_time.time().hour() < 12 else 1)  # 0 for AM, 1 for PM
            self.update_coptic_date_label()
        except Exception as e:
            self.show_error_message(str(e))

    def handle_hour_change(self, value):
        """Handle the hour spin box value change."""
        if value == 12:  # Check if the hour is 12
            # Change AM/PM based on current selection
            if self.ampm_combo.currentText() == "AM":
                self.ampm_combo.setCurrentIndex(1)  # Change to PM
            else:
                self.ampm_combo.setCurrentIndex(0)  # Change to AM

        # After reaching 12, reset to 1
        if value >= 13:  # This will cover cases for 12 and 1
            self.hour_spin.setValue(1)  # Reset to 1

        if value == 0:  # This will cover the case when value is reset to 0
            self.hour_spin.setValue(12)  # Reset to 12

        self.update_coptic_date_label()

    def handle_minute_change(self, value):
        if value >= 60:
            self.hour_spin.setValue(0)
            
        self.update_coptic_date_label()

    def submit_date_time(self):
        try:
            # Get the entered date and time
            date = self.calendar_widget.selectedDate().toString("yyyy-MM-dd")  # Get the selected date
            hour = self.hour_spin.value()
            minute = self.minute_spin.value()  # Convert to 00 or 30
            ampm = self.ampm_combo.currentText()  # Get selected AM/PM text

            time = f"{hour}:{minute:02d} {ampm}"

            if not date.strip() and not time.strip():
                self.close()
            else:
                self.date_updated.emit(date, time)
        except Exception as e:
            self.show_error_message(str(e))

    def update_date_and_time_from_coptic(self, index):
        try:
            selected_value = self.excel_dropdown.itemText(index)
            wb = load_workbook(relative_path(r"Tables.xlsx"))
            ws = wb["المناسبات"]
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=7).value == selected_value:
                    coptic_month = ws.cell(row=row, column=4).value
                    coptic_day = ws.cell(row=row, column=6).value
                    gregorian_date = self.coptic_calendar.coptic_to_gregorian([self.coptic_calendar.current_coptic_datetime[0], coptic_month, coptic_day])
                    self.calendar_widget.setSelectedDate(gregorian_date)  # Update the calendar to show this date
                    self.hour_spin.setValue(12)  # Reset to 12
                    self.minute_spin.setValue(0)  # Reset to 00
                    self.ampm_combo.setCurrentIndex(1)  # Reset to PM
                    break
            self.update_coptic_date_label()
        except Exception as e:
            self.show_error_message(str(e))

    def update_coptic_date_label(self):
        try:
            selected_date = self.calendar_widget.selectedDate()
            hour = self.hour_spin.value()
            minute = self.minute_spin.value()
            ampm = self.ampm_combo.currentText()  # Get selected AM/PM text

            # Convert hour from 12-hour format to 24-hour format
            if ampm == "PM" and hour != 12:
                hour += 12
            elif ampm == "AM" and hour == 12:
                hour = 0

            if selected_date:
                # Convert selected date to datetime object
                gregorian_date = datetime(selected_date.year(), selected_date.month(), selected_date.day(), hour, minute)
                # Convert Gregorian date to Coptic date
                coptic_date = self.coptic_calendar.gregorian_to_coptic(gregorian_date)
                coptic_date_str = f"{coptic_date[2]:02}-{self.coptic_calendar.coptic_month_name(coptic_date[1]):02}-{coptic_date[0]:04}"
                self.coptic_date_label.setText(f"التاريخ القبطي: {coptic_date_str}")
        except Exception as e:
            self.show_error_message(str(e))

    def show_error_message(self, error_message):
        QMessageBox.critical(self, "Error", error_message)
        
import sys
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QDate

if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Define initial date and time to display
    initial_date = QDate.fromString("2024-05-27", "yyyy-MM-dd")  # Today's date
    initial_time = "12:00 PM"    # Example initial time

    # Create and show ChangeDate dialog
    change_date_dialog = ChangeDate(initial_date, initial_time)
    
    # Optional: Connect to handle the signal
    def handle_date_updated(date, time):
        print(f"Selected date: {date}, time: {time}")
    
    change_date_dialog.date_updated.connect(handle_date_updated)
    
    # Show the dialog as modal
    change_date_dialog.exec_()

    sys.exit(app.exec_())