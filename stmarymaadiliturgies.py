from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QPushButton, QFrame
from PyQt5.QtGui import QPixmap, QFont, QIcon, QColor
from PyQt5.QtCore import Qt, pyqtSignal, QSize, QTimer
from PyQt5.QtWidgets import QGraphicsDropShadowEffect, QDialog
from copticDate import CopticCalendar
from datetime import datetime
from bibleWindow import bibleWindow
from NotificationBar import NotificationBar
import logging
import os
import tempfile
import traceback
from commonFunctions import (
    relative_path,
    load_background_image,
    open_presentation_relative_path,
    get_open_presentations,
    close_presentation_safe,
    close_all_presentations_safe,
)
from sys import exit, argv
from SplashScreen import ModernSplashScreen
from UpdatePrompt import UpdatePrompt
import qtawesome as qta

logger = logging.getLogger(__name__)
if not logger.handlers:
    log_path = os.path.join(tempfile.gettempdir(), "stmarymaadiliturgies.log")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s"))
    logger.addHandler(file_handler)
    logger.setLevel(logging.INFO)

class ClickableFrame(QFrame):
    clicked = pyqtSignal()

    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.notification_bar = NotificationBar(self)
        self.notification_bar.setGeometry(0, 70, self.width(), 50)
        try:
            self.current_date = datetime.now()
            self.coptic_date = CopticCalendar().gregorian_to_coptic(self.current_date)
            self.checkCopticYear(self.coptic_date[0])
            from Season import get_season
            self.season = get_season(self.current_date)
            self.bishop_window = None
            self.bishop = False
            self.GuestBishop = 0
            self.seneksar = 1
            self.setWindowTitle("St. Mary Maadi Liturgies")
            self.setWindowIcon(QIcon(relative_path(r"Data\الصور\Logo.ico")))
            self.setGeometry(400, 100, 625, 600)
            self.setFixedSize(625, 600)
            self.show_update_button = False
            self.glow_effect_counter = 0
            self.active_presentation_source = None  # Track which button opened a shared presentation
            
            # Operation flags to prevent concurrent execution
            self.operation_in_progress = False
            self.update_in_progress = False
            self._modal_dialog_open = False
            self._rebuilding_main_frame = False
            
            # Cursor management to prevent stuck loading cursor
            self.cursor_override_count = 0
            self.refresh_timer = QTimer(self)
            self.refresh_timer.setSingleShot(True)
            self.refresh_timer.timeout.connect(self.refresh_button_states)

            # Try checking for updates early
            update_found, version = self.check_for_updates_silent()
            if update_found:
                self.show_update_button = True
                self.glow_effect_counter = 1

            # Background label
            self.background_label = QLabel(self)
            self.background_label.setGeometry(0, 0, self.width(), self.height())
            try:
                load_background_image(self.background_label)
            except Exception as e:
                self._log_exception("Background image failed to load", e)
                self._safe_show_message(f"خطأ في تحميل الخلفية: {str(e)}")

            frame0 = QFrame(self)
            frame0.setGeometry(0, 0, 625, 80)
            image_label = QLabel(frame0)
            image_label.setGeometry(0, 0, 625, 80)
            image_path = relative_path(r"Data\الصور\Untitled-4.png")
            pixmap = QPixmap(image_path)
            image_label.setPixmap(pixmap)

            frame1 = ClickableFrame(self)
            frame1.setGeometry(20, 80, 585, 190)
            frame1.clicked.connect(lambda: self.open_new_window())

            label1 = QLabel(self)
            label1.setObjectName("label1")
            label1.setAlignment(Qt.AlignCenter)
            label1.setGeometry(130, 0, 455, 190)
            label1.setParent(frame1)
            font = QFont()
            font.setPointSize(30)
            font.setFamily("Calibri")
            label1.setFont(font)
            label1.setStyleSheet("color: white;")

            self.image_label = QLabel(frame1)
            self.image_label.setGeometry(0, 0, 130, 190)
            self.image_label.setScaledContents(True)

            self.frame2 = QFrame(self)
            self.refresh_button_states()
            self.restore_main_frame()

            # Create the update button (single button for both states)
            self.update_button = self.create_update_button(560)
            self.create_button("تحديث الملفات", 560, self.update_section_names)
            self.create_button("اضافة تعديل خاص", 560, lambda: self.open_confirmation_window("تعديل خاص"))
            self.create_button("إعادة تشغيل", 560, self.restart_app)

            QTimer.singleShot(0, self.update_labels)

            # Add NotificationBar
            self.notification_bar = NotificationBar(self)
            self.notification_bar.setGeometry(0, 70, self.width(), 50)

            # Frame styling (unchanged)
            frame1.setStyleSheet("""
                QFrame { 
                    background: qlineargradient(
                        x1: 0, y1: 0, x2: 1, y2: 1,
                        stop: 0 rgba(15, 46, 71, 70),
                        stop: 0.6 rgba(30, 91, 138, 70),
                        stop: 1 rgba(140, 217, 255, 50)
                    );
                    border-radius: 10px;
                    border: none;
                }
            """)
            self.frame2.setStyleSheet("""
                QFrame { 
                    background: qlineargradient(
                        x1: 0, y1: 0, x2: 1, y2: 1,
                        stop: 0 rgba(15, 46, 71, 70),
                        stop: 0.6 rgba(30, 91, 138, 70),
                        stop: 1 rgba(140, 217, 255, 50)
                    );
                    border-radius: 10px;
                    border: black 2px solid;
                }
            """)

            for frame in [frame1, self.frame2]:
                shadow = QGraphicsDropShadowEffect()
                shadow.setBlurRadius(20)
                shadow.setOffset(0)
                shadow.setColor(QColor(0, 0, 0, 100))
                frame.setGraphicsEffect(shadow)

            # Initialize PowerPoint tracking
            self.last_open_presentations = set()
            self.setup_powerpoint_event_listener()

        except Exception as e:
            stack_trace = traceback.format_exc()
            self._log_exception("Initialization failed", e)
            self._safe_show_message(f"Error: {str(e)}\n\nStack Trace:\n{stack_trace}", duration=10000)
            print(f"Initialization Error: {str(e)}\n{stack_trace}")

    def _log_exception(self, message, error):
        logger.exception("%s: %s", message, error)

    def _safe_show_message(self, message, duration=3000):
        notification_bar = getattr(self, "notification_bar", None)
        if notification_bar is not None:
            try:
                notification_bar.show_message(message, duration=duration)
                return
            except Exception as error:
                logger.exception("Notification bar failed while showing message: %s", error)
        logger.warning("Notification fallback: %s", message)
        print(message)

    def set_busy_cursor(self):
        """Safely set busy cursor with tracking to prevent stuck cursors."""
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            self.cursor_override_count += 1
        except Exception as e:
            self._log_exception("Error setting cursor", e)

    def restore_normal_cursor(self):
        """Safely restore normal cursor with tracking."""
        try:
            if self.cursor_override_count > 0:
                QApplication.restoreOverrideCursor()
                self.cursor_override_count -= 1
        except Exception as e:
            self._log_exception("Error restoring cursor", e)

    def _begin_modal_dialog(self, busy_message="عملية جارية... يرجى الانتظار"):
        if getattr(self, "_modal_dialog_open", False):
            self.notification_bar.show_message(busy_message, duration=2000)
            return False
        self._modal_dialog_open = True
        return True

    def _end_modal_dialog(self):
        self._modal_dialog_open = False

    def ensure_normal_cursor(self):
        """Force restore normal cursor by clearing all overrides."""
        try:
            while self.cursor_override_count > 0:
                QApplication.restoreOverrideCursor()
                self.cursor_override_count -= 1
            # Safety: restore even if count is wrong
            QApplication.restoreOverrideCursor()
        except Exception:
            pass

    def create_update_button(self, y):
        """Create the single update button with initial state."""
        button_texts = ["تحديث الملفات", "اضافة تعديل خاص", "تحديث", "إعادة تشغيل"]
        button_width = 115
        spacing = 10
        total_width = (button_width * len(button_texts)) + (spacing * (len(button_texts) - 1))
        start_x = (self.width() - total_width) / 2
        button_index = button_texts.index("تحديث")
        button_x = start_x + (button_index * (button_width + spacing))

        button = QPushButton(self)
        button.setGeometry(int(button_x), y, button_width, 30)

        font = QFont()
        font.setBold(True)
        font.setPointSize(9 if self.show_update_button else 8)
        button.setFont(font)

        # Set initial state based on self.show_update_button
        if self.show_update_button:
            button.setText("تحديث البرنامج")
            button.setToolTip("تحديث إلى أحدث إصدار من البرنامج")
            button.setIcon(qta.icon('fa5s.download', color='white'))
            button.clicked.connect(self.handle_update_prompt)
            if self.glow_effect_counter > 0:
                glow = QGraphicsDropShadowEffect(button)
                glow.setOffset(0)
                glow.setBlurRadius(30)
                glow.setColor(QColor(0, 255, 0))
                button.setGraphicsEffect(glow)
        else:
            button.setText("البحث عن تحديث")
            button.setToolTip("التحقق من وجود تحديث")
            button.setIcon(qta.icon('fa5s.sync-alt', color='white'))
            button.clicked.connect(self.check_for_updates_active)

        button.setIconSize(QSize(20 if self.show_update_button else 18, 20 if self.show_update_button else 18))
        button.setLayoutDirection(Qt.RightToLeft)
        button.setStyleSheet("""
            QPushButton {
                background-color: #1e5b8a;
                color: white;
                border-radius: 15px;
                font-weight: bold;
                padding: 3px;
                border: none;
            }
            QPushButton:hover {
                background-color: #3498db;
                color: white;
            }
            QPushButton:pressed {
                background-color: #2980b9;
                color: white;
            }
        """)
        return button

    def check_for_updates_active(self):
        """Check for updates and update the single button's state."""
        # Prevent concurrent update checks
        if self.update_in_progress:
            return
        
        self.update_in_progress = True
        self.update_button.setEnabled(False)
        self.set_busy_cursor()
        
        try:
            found, server_version = self.check_for_updates_silent()

            if found:
                self.show_update_button = True
                self.glow_effect_counter = 1

                # Update the existing button
                self.update_button.setText("تحديث البرنامج")
                self.update_button.setToolTip("تحديث إلى أحدث إصدار من البرنامج")
                self.update_button.setIcon(qta.icon('fa5s.download', color='white'))
                self.update_button.setIconSize(QSize(20, 20))
                font = QFont()
                font.setBold(True)
                font.setPointSize(9)
                self.update_button.setFont(font)

                # Disconnect previous signal and connect to new handler
                try:
                    self.update_button.clicked.disconnect()
                except Exception:
                    pass
                self.update_button.clicked.connect(self.handle_update_prompt)

                # Add glow effect
                glow = QGraphicsDropShadowEffect(self.update_button)
                glow.setOffset(0)
                glow.setBlurRadius(30)
                glow.setColor(QColor(0, 255, 0))
                self.update_button.setGraphicsEffect(glow)

                self.notification_bar.show_message(f"✅ تحديث جديد متوفر (الإصدار {server_version})!", duration=5000)
            else:
                # Revert to "Check for Updates" if not already in that state
                if self.show_update_button:
                    self.show_update_button = False
                    self.glow_effect_counter = 0
                    self.update_button.setText("البحث عن تحديث")
                    self.update_button.setToolTip("التحقق من وجود تحديث")
                    self.update_button.setIcon(qta.icon('fa5s.sync-alt', color='white'))
                    self.update_button.setIconSize(QSize(18, 18))
                    font = QFont()
                    font.setBold(True)
                    font.setPointSize(8)
                    self.update_button.setFont(font)
                    self.update_button.setGraphicsEffect(None)
                    try:
                        self.update_button.clicked.disconnect()
                    except Exception:
                        pass
                    self.update_button.clicked.connect(self.check_for_updates_active)

                self.notification_bar.show_message("أنت تستخدم أحدث إصدار أو لا يوجد اتصال.", duration=4000)

            # Refresh the UI
            self.update_button.show()
            self.update()
            self.repaint()
        
        finally:
            self.restore_normal_cursor()
            self.update_button.setEnabled(True)
            self.update_in_progress = False

    def add_button_with_image(self, parent, image_path, geometry, text, action=None):
        import os
        x, y, width, height = geometry

        # Create a container frame for the button with rounded corners
        container = QFrame(parent)
        container.setGeometry(x, y, width, height + 20)
        container.setStyleSheet("""
            QFrame {
                background-color: rgba(31, 91, 138, 70);
                border-radius: 10px;
            }
        """)
        
        # Check if this button corresponds to an open presentation
        open_presentations = get_open_presentations()
        presentation_map = {
            "القداس": r"قداس.pptx",
            "قداس الطفل": r"قداس الطفل.pptx",
            "باكر": r"باكر.pptx",
            "عشية": r"رفع بخور عشية و باكر.pptx",
            "الإبصلمودية": r"الإبصلمودية.pptx",
            "المدائح": r"كتاب المدائح.pptx"
        }
        
        is_open = False
        if text in presentation_map:
            full_path = os.path.abspath(relative_path(presentation_map[text])).lower()
            is_open = any(open_pres.lower() == full_path for open_pres in open_presentations)
            
            if is_open:
                # Add glow effect to container if presentation is open
                glow = QGraphicsDropShadowEffect(container)
                glow.setOffset(0)
                glow.setBlurRadius(30)
                glow.setColor(QColor(0, 255, 0))
                container.setGraphicsEffect(glow)
        
        # Image Label - Use the full width and height available
        image_label = QLabel(container)
        image_label.setGeometry(5, 5, width - 10, height - 10)
        image_label.setStyleSheet("background: transparent; border-radius: 10px;")
        
        try:
            # Load and prepare image with proper sizing
            pixmap = QPixmap(relative_path(image_path))
            if not pixmap.isNull():
                # Scale pixmap to fill the entire label
                pixmap = pixmap.scaled(width - 10, height - 10, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                
                # Center the image in the label
                image_label.setScaledContents(True)
                image_label.setPixmap(pixmap)

        except Exception as e:
            print(f"Error loading image {image_path}: {str(e)}")
            # Create a placeholder with text if image loading fails
            placeholder = QPixmap(width - 10, height - 10)
            placeholder.fill(QColor(60, 120, 190))
            image_label.setPixmap(placeholder)
        
        # Position the text label closer to the image
        label = QLabel(text, container)
        label.setAlignment(Qt.AlignCenter)
        # Move the label up to be closer to the image (reduced spacing)
        label.setGeometry(0, height-10, width, 30)
        
        # Apply font size based on text content
        font = QFont()
        if text == "الكتاب المقدس":
            font.setPointSize(10)
        else:
            font.setPointSize(12)
        font.setBold(True)
        label.setFont(font)
        label.setStyleSheet("background-color: transparent; color: white; border: none; font-weight: bold;")
        
        # Button with improved hover effect
        button = QPushButton(container)
        button.setGeometry(0, 0, width, height + 30)
        button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 25);
                border-radius: 10px;
            }
            QPushButton:pressed {
                background-color: rgba(255, 255, 255, 50);
            }
        """)
        
        if text == "صلاة السجدة":
            button.clicked.connect(lambda _, p=action: self.handle_sagda_button_click())
        else:
            button.clicked.connect(action)

    def create_button(self, text, y, action):
        button_texts = ["تحديث الملفات", "اضافة تعديل خاص"]
        if self.show_update_button:
            button_texts.append("تحديث البرنامج")
        else:
            button_texts.append("البحث عن تحديث")
        button_texts.append("إعادة تشغيل")

        button_width = 115
        spacing = 10
        total_width = (button_width * len(button_texts)) + (spacing * (len(button_texts) - 1))
        start_x = (self.width() - total_width) / 2

        try:
            button_index = button_texts.index(text)
        except ValueError:
            button_index = 0

        button_x = start_x + (button_index * (button_width + spacing))

        button = QPushButton(text, self)
        button.setGeometry(int(button_x), y, button_width, 30)
        # Font sizing
        font = QFont()
        font.setBold(True)
        font_size = 14
        if len(text) > 10:
            font_size = 9
        if len(text) > 14:
            font_size = 8
        font.setPointSize(font_size)
        button.setFont(font)

        # Default style
        button.setStyleSheet(f"""
            QPushButton {{
                background-color: #1e5b8a;
                color: white;
                border-radius: 15px;
                font-weight: bold;
                padding: 3px;
                border: none;
                font-size: {font_size}pt;
            }}
            QPushButton:hover {{
                background-color: #3498db;
                color: white;
            }}
            QPushButton:pressed {{
                background-color: #2980b9;
                color: white;
            }}
        """)

        # Special buttons
        if text == "تحديث البرنامج":
            import qtawesome as qta
            button.setToolTip("تحديث إلى أحدث إصدار من البرنامج")
            button.setLayoutDirection(Qt.RightToLeft)
            button.setIcon(qta.icon('fa5s.download', color='white'))
            button.setIconSize(QSize(20, 20))
            if self.glow_effect_counter > 0:
                glow = QGraphicsDropShadowEffect(button)
                glow.setOffset(0)
                glow.setBlurRadius(30)
                glow.setColor(QColor(0, 255, 0))
                button.setGraphicsEffect(glow)

        elif text == "البحث عن تحديث":
            import qtawesome as qta
            button.setToolTip("التحقق من وجود تحديث")
            button.setLayoutDirection(Qt.RightToLeft)
            button.setIcon(qta.icon('fa5s.sync-alt', color='white'))
            button.setIconSize(QSize(18, 18))

        elif text == "إعادة تشغيل":
            button.setToolTip("إعادة تشغيل البرنامج")
            button.setStyleSheet(f"""
                QPushButton {{
                    background-color: #e74c3c;
                    color: white;
                    border-radius: 15px;
                    font-weight: bold;
                    padding: 3px;
                    border: none;
                    font-size: {font_size}pt;
                }}
                QPushButton:hover {{
                    background-color: #c0392b;
                    color: white;
                }}
                QPushButton:pressed {{
                    background-color: #a93226;
                    color: white;
                }}
            """)

        button.clicked.connect(action)

    def open_confirmation_window(self, prayer_type="قداس"):
        from Confirmation_Dialog import Confirm
        if not self._begin_modal_dialog():
            return False
        self.bishop = False
        self.GuestBishop = 0
        self.seneksar = 1  # Default to 1 (عناوين فقط)
        try:
            # Convert coptic date list to readable string
            coptic_date_string = self.get_coptic_date_string()
            
            # Create dialog with parent and show it modally
            confirmation_dialog = Confirm(parent=self, coptic_date=coptic_date_string, type=prayer_type)
            confirmation_dialog.row2.line_edit.textChanged.connect(lambda: self.update_checkbox_state(confirmation_dialog))
            confirmation_dialog.update_button.clicked.connect(lambda: self.update_bishop_variables(confirmation_dialog))
            if confirmation_dialog.synaxar_section:
                confirmation_dialog.synaxar_section.radio1.toggled.connect(lambda: self.update_synaxar_option(confirmation_dialog))
                confirmation_dialog.synaxar_section.radio2.toggled.connect(lambda: self.update_synaxar_option(confirmation_dialog))
                # Set initial value from the dialog
                self.seneksar = confirmation_dialog.synaxar_section.get_selected_option()
            # Show as modal dialog
            result = confirmation_dialog.exec_()
            
            # Return True if user saved (accepted), False if cancelled/closed
            return result == confirmation_dialog.Accepted
        finally:
            self._end_modal_dialog()

    def get_coptic_date_string(self):
        """Convert coptic date list to readable Arabic string"""
        if not hasattr(self, 'coptic_date') or not self.coptic_date:
            return ""
        
        m = self.getmonth(self.coptic_date[1])
        m = self.convert_to_arabic_digits(m)
        coptic_date_text = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
        
        return coptic_date_text

    def update_checkbox_state(self, dialog):
        # If row2's line edit has text, check the checkbox
        if dialog.row2.line_edit.text():
            dialog.checkbox1.setChecked(True)
        else:
            dialog.checkbox1.setChecked(False)

    def update_bishop_variables(self, dialog):
        # Update self.bishop based on the checkbox state in Bishop dialog
        self.bishop = dialog.checkbox1.isChecked()

        if dialog.row2.line_edit.text():
            self.GuestBishop += 1
        if dialog.row3.line_edit.text():
            self.GuestBishop += 1
        
        # Update synaxar option if available
        if dialog.synaxar_section:
            self.seneksar = dialog.synaxar_section.get_selected_option()
        else:
            # Ensure default value if no synaxar section
            self.seneksar = 1
        
        # Close the Bishop dialog after updating variables
        dialog.accept()

    def update_synaxar_option(self, dialog):
        """Update the synaxar option when radio buttons change"""
        if dialog.synaxar_section:
            self.seneksar = dialog.synaxar_section.get_selected_option()

    def open_elmonasbat_Window(self):
        if self._rebuilding_main_frame:
            return

        self._rebuilding_main_frame = True
        try:
            old_frame = getattr(self, "frame2", None)
            if old_frame is not None:
                old_frame.deleteLater()

            # Create new frame
            self.frame2 = QFrame(self)
            self.frame2.setGeometry(20, 286, 585, 275)
            self.frame2.setStyleSheet("""
                QFrame { 
                    background: qlineargradient(
                        x1: 0, y1: 0, x2: 1, y2: 1,
                        stop: 0 rgba(15, 46, 71, 70),
                        stop: 0.6 rgba(30, 91, 138, 70),
                        stop: 1 rgba(140, 217, 255, 50)
                    );
                    border-radius: 10px;
                    border: black 2px solid;
                }
            """)

            # Animated fade-in effect
            self.fade_in_widget(self.frame2)

            # Buttons with enhanced layout
            buttons = [
                ("Data/الصور/البصخة.jpg", (13, 15, 100, 100), "اسبوع الالام", self.handle_elbas5a_button_click),
                ("Data/الصور/السجدة.jpg", (126, 15, 100, 100), "صلاة السجدة", "Data/صلاة السجدة عيد العنصرة.pptx"),
                ("Data/الصور/اللقان.jpg", (239, 15, 100, 100), "اللقان", self.handle_ellakan_button_click),
            ]

            for img, geo, label, action in buttons:
                self.add_button_with_image(self.frame2, img, geo, label, action)

            # Styled back button
            self.add_back_button(self.frame2, self.restore_main_frame)
            self.frame2.show()
        finally:
            self._rebuilding_main_frame = False

    def open_bible_window(self):
        if self.centralWidget():
            self.clear_central_widget()
        
        bible_content = bibleWindow()
        self.setCentralWidget(bible_content)

    def open_elfhrs_window(self):
        try:
            from elfhrsNEWindow import elfhrswindow

            if not self._begin_modal_dialog():
                return
            
            try:
                # Clear central widget if it exists
                if self.centralWidget():
                    self.clear_central_widget()
                
                # Create the elfhrs window content as a widget
                elfhrs_content = elfhrswindow(self)
                
                # Set it as central widget (keeps it within the main application window)
                self.setCentralWidget(elfhrs_content)
                
                # Optional: Update the UI to reflect the current state
                self.refresh_button_states(skip_timer=True)
            finally:
                self._end_modal_dialog()
            
        except Exception as e:
            import traceback
            stack_trace = traceback.format_exc()
            self.notification_bar.show_message(f"Error opening Elfhrs window: {str(e)}\n\nStack Trace:\n{stack_trace}", duration=10000)
            print(f"Elfhrs Window Error: {str(e)}\n{stack_trace}")

    def open_taranym_window(self):
        from TaranymWindow import Taranymwindow
        if not self._begin_modal_dialog():
            return
        try:
            if self.centralWidget():
                self.clear_central_widget()

            elfhrs_content = Taranymwindow()
            self.setCentralWidget(elfhrs_content)
        finally:
            self._end_modal_dialog()

    def update_section_names(self):
        from sectionNames import extract_section_info2
        
        # Prevent concurrent updates
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()
        self.notification_bar.show_message("جاري تحديث الملفات... قد يستغرق بضع ثوان", duration=3000)
        
        try:
            file_sheet_pairs = [
                (relative_path(r"Data\CopyData\قداس.pptx"), "القداس"),
                (relative_path(r"Data\CopyData\قداس الطفل.pptx"), "قداس الطفل"),
                (relative_path(r"Data\CopyData\رفع بخور عشية و باكر.pptx"), "رفع بخور"),
                (relative_path(r"Data\CopyData\الذكصولوجيات.pptx"), "الذكصولوجيات"),
                (relative_path(r"Data\CopyData\في حضور الاسقف و اساقفة ضيوف.pptx"), "في حضور الأسقف"),
                (relative_path(r"Data\CopyData\الإبصلمودية.pptx"), "التسبحة"),
                (relative_path(r"Data\CopyData\الإبصلمودية الكيهكية.pptx"), "تسبحة كيهك"),
                (relative_path(r"Data\CopyData\كتاب المدائح.pptx"), "المدائح"),
                (relative_path(r"Data\CopyData\صلاة اللقان.pptx"), "اللقان"),
                (relative_path(r"Data\اسبوع الالام\البصخة المقدسة.pptx"), "البصخة"),
                (relative_path(r"Data\اسبوع الالام\خميس العهد.pptx"), "خميس العهد"),
                (relative_path(r"Data\اسبوع الالام\الجمعة العظيمة.pptx"), "الجمعة العظيمة"),
                (relative_path(r"Data\CopyData\صلاة السجدة.pptx"), "صلاة السجدة"),
                ]

            excel_file = relative_path(r'Files Data.xlsx')
            
            extract_section_info2(file_sheet_pairs, excel_file)

            # Show success message
            self.show_message("تم التحديث بنجاح!")

        except Exception as e:
            self.show_error_message(str(e))
        finally:
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def season_picture(self):
        match self.season :
            case 0:
                return r"Data\الصور\Aykona.png"
            case 1 | 1.1:
                return r"Data\الصور\النيروز.JPG"
            case 4 | 4.1:
                return r"Data\الصور\عيد الميلاد.jpg"
            case 10 :
                return r"Data\الصور\عرس قانا الجليل.jpg"
            case 17:
                return r"Data\الصور\الشعانين.jpg"
            case 19:
                return r"Data\الصور\خميس العهد.jpg"
            case 20 | 18:
                return r"Data\الصور\الجمعة العظيمةو البصخة.jpg"
            case 21:
                return r"Data\الصور\سبت النور.JPG"
            case 22 | 24:
                return r"Data\الصور\القيامة.jpg"
            case 23.3 | 24.1 | 25:
                return r"Data\الصور\الصعود.jpg"
            case 23.1 | 23:
                return r"Data\الصور\دخول المسيح أرض مصر.jpg"
            case 29 :
                return r"Data\الصور\التجلي.JPG"
        return r"Data\الصور\Aykona.png" 

    def open_new_window(self):
        from ChangeDateWindow import ChangeDate
        if not self._begin_modal_dialog():
            return
        try:
            new_window = ChangeDate(self.current_date.date(), self.current_date.strftime("%I:%M %p"))
            new_window.date_updated.connect(self.update_current_date)
            new_window.exec_()
        finally:
            self._end_modal_dialog()

    def clear_central_widget(self):
        central_widget = self.centralWidget()
        if central_widget:
            layout = central_widget.layout()
            if layout:
                while layout.count():
                    child = layout.takeAt(0)
                    if child.widget():
                        child.widget().deleteLater()
                self.setCentralWidget(None)

    def update_current_date(self, new_date, new_time):
        from Season import get_season
        try:
            self.current_date = datetime.strptime(new_date + ' ' + new_time, '%Y-%m-%d %I:%M %p')
            self.coptic_date = CopticCalendar().gregorian_to_coptic(self.current_date)
            self.season = get_season(self.current_date)
            QTimer.singleShot(0, self.update_labels)
            self.close_dialog()
        except ValueError:
            self.show_error_message("التاريخ/الوقت غير صحيح.")

    def convert_to_arabic_digits(self, number):
        arabic_digits = {'0': '٠', '1': '١', '2': '٢', '3': '٣', '4': '٤', '5': '٥', '6': '٦', '7': '٧', '8': '٨', '9': '٩'}
        return ''.join(arabic_digits[digit] if digit in arabic_digits else digit for digit in str(number))

    def update_labels(self):
        from Season import get_season_name
        from PyQt5.QtGui import QFontMetrics
        label1 = self.findChild(QLabel, "label1")
        if label1:
            sesn = get_season_name(self.season)
            m = self.getmonth(self.coptic_date[1])
            m = self.convert_to_arabic_digits(m)
            ad = self.get_arabic_month_date(self.current_date)
            ad = self.convert_to_arabic_digits(ad)
            c = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
            if self.current_date.time() > datetime.strptime('5:30 PM', '%I:%M %p').time():
                c = f"({c})"
            
            # Create a QFontMetrics object to measure text width
            font = QFont()
            font.setPointSize(30)
            font.setFamily("Calibri")
            
            # Format the complete text to measure the full content
            date_text = f"{sesn}\n{c}\n{ad}"
            
            # Calculate available space
            max_width = label1.width() - 20  # 10px padding on each side
            max_height = label1.height() - 20  # 10px padding on top and bottom
            
            # Measure text dimensions with current font
            font_metrics = QFontMetrics(font)
            text_rect = font_metrics.boundingRect(0, 0, max_width, 1000, 
                                            Qt.AlignCenter | Qt.TextWordWrap, 
                                            date_text)
            
            text_width = font_metrics.width(sesn)  # Check if season name fits in one line
            text_height = text_rect.height()
            
            # If text exceeds width or height, reduce font size
            if text_width > max_width or text_height > max_height:
                # Gradually decrease font size until text fits or min size reached
                adjusted_size = 30  # Start with default
                
                while adjusted_size > 18 and (text_width > max_width or text_height > max_height):
                    adjusted_size -= 2
                    font.setPointSize(adjusted_size)
                    font_metrics = QFontMetrics(font)
                    
                    # Recalculate dimensions with new font size
                    text_width = font_metrics.width(sesn)
                    text_rect = font_metrics.boundingRect(0, 0, max_width, 1000, 
                                                    Qt.AlignCenter | Qt.TextWordWrap, 
                                                    date_text)
                    text_height = text_rect.height()
                
                # Update the label's font
                label1.setFont(font)
            else:
                # If it fits, use the original size
                font.setPointSize(30)
                label1.setFont(font)
            
            # Set the text
            label1.setText(date_text)
        
        new_pixmap = QPixmap(relative_path(self.season_picture()))
        self.image_label.setPixmap(new_pixmap)

    def close_dialog(self):
        from ChangeDateWindow import ChangeDate
        for widget in QApplication.instance().topLevelWidgets():
            if isinstance(widget, ChangeDate):
                widget.close()

    def show_error_message(self, error_message):
        import traceback
        stack_trace = traceback.format_exc()
        full_error = f"Error: {error_message}\n\nStack Trace:\n{stack_trace}"
        self._safe_show_message(full_error, duration=10000)  # Longer duration for stack traces
        print(full_error)  # Also print to console for debugging

    def show_message(self, message):
        self._safe_show_message(message, duration=3000)

    def handle_qadas_button_click(self):
        import odasat
        from qudasDialog import SectionSelectionDialog
        import os
        from Season import get_season_name
        from PyQt5.QtWidgets import QMessageBox
        
        # Prevent concurrent operations
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()

        try:
            presentation_opened = False
            presentation_file = os.path.abspath(relative_path(r"قداس.pptx")).lower()
            
            # Check if presentation is already open
            open_presentations = get_open_presentations()
            is_already_open = any(open_pres.lower() == presentation_file for open_pres in open_presentations)
            
            # Check if current season has a qudas function available
            # Seasons that don't have qudas functions implemented yet
            seasons_without_function = [
                11,   # دخول المسيح الهيكل
                18,   # أسبوع الالام
                19,   # خميس العهد
                20,   # الجمعة العظيمة
                21,   # سبت النور
                23,   # دخول المسيح أرض مصر
                23.1, # دخول المسيح أرض مصر والخمسين المقدسة
                23.2, # عيد العنصرة ودخول المسيح أرض مصر
                23.3  # عيد الصعود ودخول المسيح أرض مصر
            ]
            
            if self.season in seasons_without_function:
                # Show appropriate message based on season type
                if self.season in [19, 20, 21]:
                    # Special messages for Holy Week
                    match self.season:
                        case 19:
                            self.notification_bar.show_message("صلوات خميس العهد متوفرة في ملف واحد: المناسبات > اسبوع الالام > خميس العهد", 10000)
                        case 20:
                            self.notification_bar.show_message("لا يوجد قداس يوم الجمعة العظيمة: المناسبات > اسبوع الالام > الجمعة العظيمة", 10000)
                        case 21:
                            self.notification_bar.show_message("صلوات سبت الفرح متوفرة في ملف واحد: المناسبات > اسبوع الالام > ليلة ابوغلمسيس", 10000)
                else:
                    # Generic message for other unimplemented seasons
                    self.notification_bar.show_message(f"قداس {get_season_name(self.season)} غير متوفر حاليا", duration=5000)
                return
            
            if is_already_open:
                # If already open, just show the dialog without reopening the presentation
                presentation_opened = True
            else:
                # Restore normal cursor before showing dialog to user
                self.restore_normal_cursor()
                
                # Show confirmation dialog for bishop settings
                result = self.open_confirmation_window("قداس")
                
                if not result:
                    return  # User cancelled/closed dialog, exit the function
                
                # User confirmed and saved settings, set busy cursor again for processing
                self.set_busy_cursor()
                
                # Proceed with opening the presentation
                match self.season:
                    case 0 | 6 | 13 | 30 | 31:
                        odasat.odasSanawy(self.coptic_date, self.season, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 1 | 1.1:
                        odasat.odasElnayrooz(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                    case 2:
                        odasat.odasElsalyb(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 3 | 8:
                        odasat.odasbaramonElmiladAndEl8etas(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 4:
                        odasat.odasElmilad(self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 4.1 | 4.2:
                        odasat.odasAfterElmilad(self.coptic_date, self.bishop, self.GuestBishop)
                    case 5:
                        odasat.odasKiahk(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 7:
                        odasat.odasEl5etan(self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 9 | 9.1:
                        odasat.odasEl8ytas(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 10:
                        odasat.odas3orsKanaElgalyl(self.bishop, self.GuestBishop, self.seneksar)                        
                        presentation_opened = True
                    case 12:
                        odasat.odasSomNynawa(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 14:
                        odasat.odasElbeshara(self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 15 | 15.1 | 15.2 | 15.3 | 15.4 | 15.5 | 15.6 | 15.7 | 15.8 | 15.9 | 15.11:
                        odasat.odasElSomElkbyr(self.coptic_date, self.season, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 16:
                        odasat.odasSbtLe3azr(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 17:
                        odasat.odasElsh3anyn(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 22:
                        odasat.odasEl2yama(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 24:
                        odasat.odasEl5amasyn_2_39(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 24.1:
                        odasat.odasElso3od(self.coptic_date, self.bishop, self.GuestBishop, True)
                        presentation_opened = True
                    case 25:
                        odasat.odasElso3od(self.coptic_date, self.bishop, self.GuestBishop, False)
                        presentation_opened = True
                    case 26:
                        odasat.odasEl3nsara(self.coptic_date, self.bishop, self.GuestBishop)
                        presentation_opened = True
                    case 27:
                        odasat.odasSomElRosol(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 28:
                        odasat.odas3ydElrosol(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 29:
                        odasat.odasEltagaly(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case 32:
                        odasat.odas29thOfMonth(self.coptic_date, self.bishop, self.GuestBishop, self.seneksar)
                        presentation_opened = True
                    case default:
                        self.notification_bar.show_message(f"قداس {get_season_name(self.season)} غير متوفر حاليا")
            
            # After the match logic, if a presentation was opened, show the sections dialog
            if presentation_opened:
                self.refresh_button_states()
                self.restore_main_frame()
                # Format dates for the dialog
                m = self.getmonth(self.coptic_date[1])
                m = self.convert_to_arabic_digits(m)
                coptic_date_text = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
                arabic_date_text = self.get_arabic_month_date(self.current_date)
                title = f"قداس {coptic_date_text} / {arabic_date_text}"
                sheet_name = "القداس"
                # Restore normal cursor before showing dialog to user
                self.restore_normal_cursor()
                # Create and show the dialog
                dialog = SectionSelectionDialog.get_dialog(self, title, sheet_name)
                dialog.exec_()
                # Set busy cursor again if needed for any post-dialog processing
                self.set_busy_cursor()
                
        except Exception as e:
            import traceback
            stack_trace = traceback.format_exc()
            self.notification_bar.show_message(f"خطأ في فتح القداس: {str(e)}", duration=5000)
            print(f"Qudas Error: {str(e)}\n{stack_trace}")
        finally:
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def handle_qadas_eltfl_button_click(self):
        # from odasatEltfl import (odasElSomElkbyr, odasEltflSomElrosol, odasEltfl3ydElrosol, odasSanawy, 
        #                          odasEltflElnayrooz, odasEltflKiahk)
        # try:
        #     if(self.pptx_check(True) == False):
        #         self.replace_presentation(True)
        #     match self.season:
        #         case 0 | 6 | 30 | 31:
        #             odasSanawy(self.coptic_date, self.season)
        #         case 1:
        #             odasEltflElnayrooz(self.coptic_date)
        #         case 5:
        #             odasEltflKiahk(self.coptic_date)
        #         case 15 | 15.1:
        #             odasElSomElkbyr(self.coptic_date, self.season)
        #         case 27:
        #             odasEltflSomElrosol(self.coptic_date)
        #         case 28:
        #             odasEltfl3ydElrosol()
        #         case default :
        #             self.notification_bar.show_message(f"قداس {get_season_name(self.season)} غير متوفر حاليا")
        # except Exception as e:
        #     self.show_error_message(str(e))
        return

    def handle_baker_button_click(self):
        from baker import bakerSanawy, bakerKiahk, bakerElmilad, bakerEl8etas
        from qudasDialog import SectionSelectionDialog
        from PyQt5.QtWidgets import QMessageBox
        import os
        from Season import get_season_name
        
        # Prevent concurrent operations
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()
        
        try:
            presentation_file = os.path.abspath(relative_path(r"رفع بخور عشية و باكر.pptx")).lower()
            
            # Check if presentation is already open
            open_presentations = get_open_presentations()
            is_already_open = any(open_pres.lower() == presentation_file for open_pres in open_presentations)
            
            if is_already_open and self.active_presentation_source != "باكر":
                # File is open but was opened by a different button
                self.restore_normal_cursor()
                reply = QMessageBox.question(self, "تحذير",
                                  "هذا الملف مفتوح حاليًا في عرض عشية. هل تريد إغلاق الملف وفتحه كباكر؟",
                                  QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                
                if reply == QMessageBox.Yes:
                    self.set_busy_cursor()
                    closed = close_presentation_safe(presentation_file)
                    if not closed:
                        self.notification_bar.show_message("تعذر إغلاق الملف الحالي", duration=3000)
                    self.active_presentation_source = None

                    # Now proceed with opening the file
                    is_already_open = False
                else:
                    # User chose not to proceed
                    return
            
            presentation_opened = False
            
            if is_already_open and self.active_presentation_source == "باكر":
                # If already open and was opened by this button, just show the dialog
                presentation_opened = True
            else:
                # Restore cursor before showing confirmation dialog to user
                self.restore_normal_cursor()
                # Show confirmation dialog for bishop settings
                result = self.open_confirmation_window("باكر")
                
                if not result:
                    return  # User cancelled/closed dialog, exit the function
                
                # User confirmed, set busy cursor for processing
                self.set_busy_cursor()
                # User confirmed and saved settings, proceed with opening the presentation
                adam = False
                if self.current_date.weekday() in [0, 1, 6]:
                    adam = True
                match self.season:
                    case 0 | 27 | 28 | 30 | 31:
                        bakerSanawy(self.season, self.coptic_date, adam, self.bishop, self.GuestBishop)
                        self.active_presentation_source = "باكر"  # Set the active button
                        presentation_opened = True
                    case 4: 
                        bakerElmilad(self.season, self.coptic_date, self.bishop, self.GuestBishop)
                    case 5:
                        bakerKiahk(self.coptic_date, adam, self.bishop, self.GuestBishop)
                        self.active_presentation_source = "باكر"  # Set the active button
                        presentation_opened = True
                    case 9 | 9.1:
                        bakerEl8etas(self.season, self.coptic_date, self.bishop, self.GuestBishop)
                        self.active_presentation_source = "باكر"  # Set the active button
                        presentation_opened = True
                    case _:
                        self.notification_bar.show_message(f"رفع بخور باكر {get_season_name(self.season)} غير متوفر حاليا")
            
            if presentation_opened:
                self.refresh_button_states(skip_timer=True)
                self.restore_main_frame()
                # Format dates for the dialog
                m = self.getmonth(self.coptic_date[1])
                m = self.convert_to_arabic_digits(m)
                coptic_date_text = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
                arabic_date_text = self.get_arabic_month_date(self.current_date)
                title = f"رفع بخور باكر {coptic_date_text} / {arabic_date_text}"
                sheet_name = "رفع بخور"
                # Restore cursor before showing dialog to user
                self.restore_normal_cursor()
                # Create and show the dialog
                dialog = SectionSelectionDialog.get_dialog(self, title, sheet_name)
                dialog.exec_()
        
        except Exception as e:
            self.notification_bar.show_message(f"خطأ في فتح باكر: {str(e)}", duration=5000)
            print(f"Baker Error: {str(e)}")
        finally:
            self.restore_normal_cursor()
            self.operation_in_progress = False    

    def handle_3ashya_button_click(self):
        from Aashya import aashyaKiahk, aashyaSanawy, aashyaEltagaly
        from qudasDialog import SectionSelectionDialog
        from PyQt5.QtWidgets import QMessageBox
        import os
        
        # Prevent concurrent operations
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()
        
        try:
            presentation_file = os.path.abspath(relative_path(r"رفع بخور عشية و باكر.pptx")).lower()
            
            # Check if presentation is already open
            open_presentations = get_open_presentations()
            is_already_open = any(open_pres.lower() == presentation_file for open_pres in open_presentations)
            
            if is_already_open and self.active_presentation_source != "عشية":
                # File is open but was opened by a different button
                self.restore_normal_cursor()
                reply = QMessageBox.question(self, "تحذير",
                              "هذا الملف مفتوح حاليًا في عرض باكر. هل تريد إغلاق الملف وفتحه كعشية؟",
                              QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                
                if reply == QMessageBox.Yes:
                    self.set_busy_cursor()
                    closed = close_presentation_safe(presentation_file)
                    if not closed:
                        self.notification_bar.show_message("تعذر إغلاق الملف الحالي", duration=3000)
                    self.active_presentation_source = None

                    # Now proceed with opening the file
                    is_already_open = False
                else:
                    # User chose not to proceed
                    return
            
            presentation_opened = False
            
            if is_already_open and self.active_presentation_source == "عشية":
                # If already open and was opened by this button, just show the dialog
                presentation_opened = True
            else:
                # Restore cursor before showing confirmation dialog to user
                self.restore_normal_cursor()
                # Show confirmation dialog for bishop settings
                result = self.open_confirmation_window("عشية")
                
                if not result:
                    return  # User cancelled/closed dialog, exit the function
                
                # User confirmed, set busy cursor for processing
                self.set_busy_cursor()
                # User confirmed and saved settings, proceed with opening the presentation
                adam = False
                if self.current_date.weekday() in [0, 1, 6]:
                    adam = True
                
                match (self.season):
                    case 0 | 27 | 30 | 31:
                        aashyaSanawy(self.season, self.coptic_date, adam, self.bishop, self.GuestBishop)
                        self.active_presentation_source = "عشية"  # Set the active button
                        presentation_opened = True
                    case 5:
                        aashyaKiahk(self.coptic_date, adam, self.bishop, self.GuestBishop)
                        self.active_presentation_source = "عشية"  # Set the active button
                        presentation_opened = True
                    case 29:
                        aashyaEltagaly(self.coptic_date, adam, self.bishop, self.GuestBishop)
                        self.active_presentation_source = "عشية"  # Set the active button
                        presentation_opened = True
            
            if presentation_opened:
                self.refresh_button_states(skip_timer=True)
                self.restore_main_frame()
                # Format dates for the dialog
                m = self.getmonth(self.coptic_date[1])
                m = self.convert_to_arabic_digits(m)
                coptic_date_text = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
                arabic_date_text = self.get_arabic_month_date(self.current_date)
                title = f"رفع بخور عشية {coptic_date_text} / {arabic_date_text}"
                sheet_name = "رفع بخور"
                # Restore cursor before showing dialog to user
                self.restore_normal_cursor()
                # Create and show the dialog
                dialog = SectionSelectionDialog.get_dialog(self, title, sheet_name)
                dialog.exec_()
                
        except Exception as e:
            import traceback
            stack_trace = traceback.format_exc()
            self.notification_bar.show_message(f"خطأ في فتح عشية: {str(e)}", duration=5000)
            print(f"Aashya Error: {str(e)}\n{stack_trace}")
        finally:
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def handle_tasbha_button_click(self):
        from tasbhaDialog import TasbhaSelectionDialog
        import tasbha
        from qudasDialog import SectionSelectionDialog
        import os
        
        # Prevent concurrent operations
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()
        
        try:
            if not self._begin_modal_dialog():
                return
            # Check if either Tasbha presentation is already open
            standard_tasbha_file = os.path.abspath(relative_path(r"الإبصلمودية.pptx")).lower()
            kiahk_tasbha_file = os.path.abspath(relative_path(r"الإبصلمودية الكيهكية.pptx")).lower()
            
            open_presentations = get_open_presentations()
            standard_is_open = any(open_pres.lower() == standard_tasbha_file for open_pres in open_presentations)
            kiahk_is_open = any(open_pres.lower() == kiahk_tasbha_file for open_pres in open_presentations)
            
            # If either tasbha file is already open, skip the tasbha selection dialog
            if standard_is_open or kiahk_is_open:
                # Format dates for the dialog
                m = self.getmonth(self.coptic_date[1])
                m = self.convert_to_arabic_digits(m)
                coptic_date_text = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
                arabic_date_text = self.get_arabic_month_date(self.current_date)
                
                # Determine which type of tasbha is open
                if kiahk_is_open:
                    title_prefix = "تسبحة كيهك"
                    sheet_name = "تسبحة كيهك"
                else:  # standard_is_open
                    title_prefix = "تسبحة"
                    sheet_name = "التسبحة"
                
                title = f"{title_prefix} {coptic_date_text} / {arabic_date_text}"
                
                # Restore cursor before showing dialog to user
                self.restore_normal_cursor()
                # Create and show the dialog
                sections_dialog = SectionSelectionDialog.get_dialog(self, title, sheet_name)
                sections_dialog.exec_()
                return
            
            # If no tasbha is open, show the selection dialog
            self.restore_normal_cursor()
            dialog = TasbhaSelectionDialog(self, self.season)
            result = dialog.exec_()

            if result == QDialog.Accepted and dialog.selected_option:
                presentation_file = None
                presentation_opened = False
                
                self.set_busy_cursor()
                # Run the corresponding tasbha function based on user selection
                if dialog.selected_option == "midnight":
                    # Run midnight tasbha
                    if self.season == 5:  # Kiahk season
                        tasbha.kiahk(self.coptic_date)
                        presentation_file = relative_path(r"الإبصلمودية الكيهكية.pptx")
                    else:
                        tasbha.tasbha(self.coptic_date, False, self.season)
                        presentation_file = relative_path(r"الإبصلمودية.pptx")
                    presentation_opened = True
                    
                elif dialog.selected_option == "evening":
                    # Run evening tasbha
                    if self.season == 5:  # Kiahk season
                        tasbha.kiahk_aashya(self.coptic_date)
                        presentation_file = relative_path(r"الإبصلمودية الكيهكية.pptx")
                    else:
                        tasbha.tasbha(self.coptic_date, True, self.season)
                        presentation_file = relative_path(r"الإبصلمودية.pptx")
                    presentation_opened = True
                
                # After opening the presentation, show the section selection dialog
                if presentation_opened and presentation_file:
                    self.refresh_button_states(skip_timer=True)
                    self.restore_main_frame()
                    
                    # Format dates for the dialog
                    m = self.getmonth(self.coptic_date[1])
                    m = self.convert_to_arabic_digits(m)
                    coptic_date_text = f"{self.convert_to_arabic_digits(self.coptic_date[2])} {m}، {self.convert_to_arabic_digits(self.coptic_date[0])}"
                    arabic_date_text = self.get_arabic_month_date(self.current_date)
                    
                    # Determine the type of tasbha for the title
                    if dialog.selected_option == "midnight":
                        title_prefix = "تسبحة نصف الليل"
                        sheet_name = "تسبحة كيهك" if self.season == 5 else "التسبحة"
                    else:
                        title_prefix = "تسبحة عشية كيهك" if self.season == 5 else "تسبحة عشية"
                        sheet_name = "تسبحة كيهك" if self.season == 5 else "التسبحة"
                    
                    title = f"{title_prefix} {coptic_date_text} / {arabic_date_text}"
                    
                    # Restore cursor before showing dialog to user
                    self.restore_normal_cursor()
                    # Create and show the dialog
                    sections_dialog = SectionSelectionDialog.get_dialog(self, title, sheet_name)
                    sections_dialog.exec_()
                    
        except Exception as e:
            import traceback
            stack_trace = traceback.format_exc()
            self.notification_bar.show_message(f"خطأ في فتح التسبحة: {str(e)}", duration=5000)
            print(f"Tasbha Error: {str(e)}\n{stack_trace}")
        finally:
            self._end_modal_dialog()
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def handle_ellakan_button_click(self):
        from elLakanDialog import LakanSelectionDialog
        import lakan
        
        # Prevent concurrent operations
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()
        
        try:
            if not self._begin_modal_dialog():
                return
            # Show the selection dialog
            dialog = LakanSelectionDialog(self)
            self.restore_normal_cursor()
            result = dialog.exec_()
            adam = False
            if self.current_date.weekday() in [0, 1, 6]:
                adam = True
            if result == QDialog.Accepted and dialog.selected_option:
                # Run the corresponding tasbha function based on user selection
                if dialog.selected_option == "Baptism":
                    lakan.lakanEl8etas(adam)
                elif dialog.selected_option == "Holy Thursday":
                    open_presentation_relative_path(r"Data\اسبوع الالام\خميس العهد.pptx")
                elif dialog.selected_option == "Apostles":
                    open_presentation_relative_path(r"Data\لقان عيد الرسل.pptx")
        except Exception as e:
            import traceback
            stack_trace = traceback.format_exc()
            self.notification_bar.show_message(f"خطأ في فتح اللقان: {str(e)}", duration=5000)
            print(f"Lakan Error: {str(e)}\n{stack_trace}")
        finally:
            self._end_modal_dialog()
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def handle_elbas5a_button_click(self):
        from elbas5aDialog import Elbas5aDialog
        
        # Prevent concurrent operations
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return
        
        self.operation_in_progress = True
        self.set_busy_cursor()
        
        try:
            if not self._begin_modal_dialog():
                return
            self.restore_normal_cursor()
            dialog = Elbas5aDialog(self)
            dialog.exec_()
        except Exception as e:
            self.notification_bar.show_message(f"خطأ في فتح أسبوع الآلام: {str(e)}", duration=5000)
            print(f"Elbas5a Error: {str(e)}")
        finally:
            self._end_modal_dialog()
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def handle_agbya_button_click(self):
        return
    
    def handle_sagda_button_click(self):
        """Open the sagda confirmation dialog first, then launch the presentation if accepted."""
        import elsagda
        if self.operation_in_progress:
            self.notification_bar.show_message("عملية جارية... يرجى الانتظار", duration=2000)
            return

        self.operation_in_progress = True

        try:
            result = self.open_confirmation_window("صلاة السجدة")
            if result:
                self.set_busy_cursor()
                elsagda.elsagda(self.season, self.bishop, self.GuestBishop)
        except Exception as e:
            self.notification_bar.show_message(f"خطأ في فتح صلاة السجدة: {str(e)}", duration=5000)
            print(f"Sagda Error: {str(e)}")
        finally:
            self.restore_normal_cursor()
            self.operation_in_progress = False

    def replace_presentation(self, odasEltfl = False, baker = False, tasbha = False, aashya = False):
        from shutil import copy2
        from os import path, remove
        if(odasEltfl):    
            old_presentation_path = relative_path(r"قداس الطفل.pptx")
            new_presentation_path = relative_path(r"Data\CopyData\قداس الطفل.pptx")
        elif(baker):
            old_presentation_path = relative_path(r"باكر.pptx")
            new_presentation_path = relative_path(r"Data\CopyData\باكر.pptx")
        elif(tasbha):
            old_presentation_path = relative_path(r"الإبصلمودية.pptx")
            new_presentation_path = relative_path(r"Data\CopyData\الإبصلمودية.pptx")
        elif(aashya):
            old_presentation_path = relative_path(r"رفع بخور عشية و باكر.pptx")
            new_presentation_path = relative_path(r"Data\CopyData\رفع بخور عشية و باكر.pptx")
        else:    
            old_presentation_path = relative_path(r"قداس.pptx")
            new_presentation_path = relative_path(r"Data\CopyData\قداس.pptx")
        try:
            # Check if the old presentation file exists
            if path.exists(old_presentation_path):
                # If it exists, delete the old presentation
                remove(old_presentation_path)
                
                # Copy the new presentation to the location of the old presentation
                copy2(new_presentation_path, old_presentation_path)
        except Exception as e:
            # Print any errors that occur during the deletion and copying process
            print(f"Error: {str(e)}")

    def get_arabic_month_date(self, current_date):
        # Define a dictionary to map month names from English to Arabic
        month_names_arabic = {
            'January': 'يناير',
            'February': 'فبراير',
            'March': 'مارس',
            'April': 'أبريل',
            'May': 'مايو',
            'June': 'يونيو',
            'July': 'يوليو',
            'August': 'أغسطس',
            'September': 'سبتمبر',
            'October': 'أكتوبر',
            'November': 'نوفمبر',
            'December': 'ديسمبر'
        }
        
        # Define a dictionary to map day names from English to Arabic
        day_names_arabic = {
            'Monday': 'الاثنين',
            'Tuesday': 'الثلاثاء',
            'Wednesday': 'الأربعاء',
            'Thursday': 'الخميس',
            'Friday': 'الجمعة',
            'Saturday': 'السبت',
            'Sunday': 'الأحد'
        }

        arabic_month = month_names_arabic[current_date.strftime('%B')]
        arabic_day = day_names_arabic[current_date.strftime('%A')]
        
        arabic_date_string = f"{arabic_day}، {current_date.day} {arabic_month} {current_date.year}"
        return arabic_date_string

    def getmonth(self, num):
        from openpyxl import load_workbook
        # Load the Excel file
        workbook = load_workbook(relative_path(r'Tables.xlsx'))
        sheet = workbook["المناسبات"]
        search_number = num 
        corresponding_value = None
        for row in sheet.iter_rows(values_only=True):
            if row[0] == search_number: 
                corresponding_value = row[1] 
                break
        return  corresponding_value

    def add_back_button(self, parent, action):
        # Get frame geometry
        frame_geometry = parent.geometry()
        # Calculate button position (bottom right corner)
        button_width = 100
        button_height = 30
        button_x = frame_geometry.width() - button_width - 10
        button_y = frame_geometry.height() - button_height - 10

        # Add back button
        back_button = QPushButton("Back", parent)
        back_button.setGeometry(button_x, button_y, button_width, button_height)
        back_button.clicked.connect(action)
        back_button.setText("⬅ العودة")
        back_button.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                font-weight: bold;
                border-radius: 12px;
                padding: 6px 14px;
                font-size: 11pt;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)

    def fade_in_widget(self, widget, duration=400): 
        from PyQt5.QtCore import QPropertyAnimation
        widget.setWindowOpacity(0)
        anim = QPropertyAnimation(widget, b"windowOpacity")
        anim.setDuration(duration)
        anim.setStartValue(0)
        anim.setEndValue(1)
        anim.start()
        widget.anim = anim  # Keep a reference so it's not garbage collected

    def restore_main_frame(self):
        if self._rebuilding_main_frame:
            return

        self._rebuilding_main_frame = True
        try:
            old_frame = getattr(self, "frame2", None)
            if old_frame is not None:
                old_frame.deleteLater()

            self.frame2 = QFrame(self)
            self.frame2.setGeometry(20, 280, 585, 275)

            self.add_button_with_image(self.frame2, "Data/الصور/القداس.JPG", (13, 15, 100, 100), "القداس", self.handle_qadas_button_click)
            self.add_button_with_image(self.frame2, "Data/الصور/قداس الطفل.png", (126, 15, 100, 100), "قداس الطفل", self.handle_qadas_eltfl_button_click)
            self.add_button_with_image(self.frame2, "Data\الصور\باكر.jpg", (239, 15, 100, 100), "باكر", self.handle_baker_button_click)
            self.add_button_with_image(self.frame2, "Data\الصور\عشية.jpg", (352, 15, 100, 100), "عشية", self.handle_3ashya_button_click)
            self.add_button_with_image(self.frame2, "Data/الصور/الكتاب المقدس.png", (465, 15, 100, 100), "الكتاب المقدس", self.open_bible_window)
            self.add_button_with_image(self.frame2, "Data\الصور\الأجبية.jpg", (13, 148, 100, 100), "الأجبية", self.handle_agbya_button_click)
            self.add_button_with_image(self.frame2, "Data\الصور\داود 1.jpg", (126, 148, 100, 100), "الإبصلمودية", self.handle_tasbha_button_click)
            self.add_button_with_image(self.frame2, "Data\الصور\الفهرس.jpg", (239, 148, 100, 100), "الفهرس", self.open_elfhrs_window)
            self.add_button_with_image(self.frame2, "Data\الصور\المدائح2.jpg", (352, 148, 100, 100), "المدائح", self.open_taranym_window)
            self.add_button_with_image(self.frame2, "Data\الصور\الصليب القبطي.jpg", (465, 148, 100, 100), "المناسبات", self.open_elmonasbat_Window)

            self.frame2.show()
        finally:
            self._rebuilding_main_frame = False

    def is_powerpoint_open(self):
        """Check if any PowerPoint application is open."""
        return len(get_open_presentations()) > 0

    def checkCopticYear(self, copticYear):
        from commonFunctions import read_excel_cell, relative_path
        import traceback
        
        try:
            currentYear = read_excel_cell(relative_path(r"Tables.xlsx"), "المناسبات", "M2")
            if copticYear != currentYear:
                # Import the unified update function
                from UpdateTable import update_all_tables
                
                # Run the update with the new year
                success = update_all_tables(copticYear)
                
                if success:
                    return True
                else:
                    return False
            else:
                return False
                
        except Exception as e:
            print(f"Error in checkCopticYear: {str(e)}")
            print(traceback.format_exc())
            return False

    def check_for_updates_silent(self):
        import socket, requests
        try:
            def have_internet_connection():
                try:
                    socket.create_connection(("8.8.8.8", 53), timeout=3)
                    return True
                except OSError:
                    return False

            if not have_internet_connection():
                return False, None

            local_version = "2.5"
            dropbox_url = "https://www.dropbox.com/scl/fi/tumjwytg8ptr88zs5pojd/version.json?rlkey=4fukyqxjx9lii0j0tunwxwpi7&st=sqk5fl08&dl=1"
            response = requests.get(dropbox_url, timeout=5)
            response.raise_for_status()
            server_version = response.json().get("version", "1.0.0")

            return (server_version > local_version), server_version

        except Exception as e:
            self._log_exception("Update check failed", e)
            return False, None

    def _pulse_glow(self, effect):
        if self._increasing:
            self._blur += 1
            if self._blur >= 45:
                self._increasing = False
        else:
            self._blur -= 1
            if self._blur <= 30:
                self._increasing = True
        effect.setBlurRadius(self._blur)

    def handle_update_prompt(self):
        import requests
        import socket

        def have_internet_connection():
            try:
                socket.create_connection(("8.8.8.8", 53), timeout=3)
                return True
            except OSError:
                return False

        if not have_internet_connection():
            self.notification_bar.show_message("⚠ لا يوجد اتصال بالإنترنت. تحقق من الاتصال وحاول مرة أخرى.", duration=5000)
            return

        try:
            if not self._begin_modal_dialog():
                return
            url = "https://www.dropbox.com/scl/fi/tumjwytg8ptr88zs5pojd/version.json?rlkey=4fukyqxjx9lii0j0tunwxwpi7&st=sqk5fl08&dl=1"
            response = requests.get(url, timeout=5)
            response.raise_for_status()
            data = response.json()

            version = data.get("version", "??")
            notes = data.get("description", "لا توجد تفاصيل.")
            exe_url = data.get("download_url")

            dialog = UpdatePrompt(version, notes, self)
            dialog.update_button.clicked.connect(lambda: self.download_update(exe_url))
            dialog.cancel_button.clicked.connect(dialog.close)
            dialog.exec_()

        except requests.exceptions.ConnectionError:
            self.notification_bar.show_message("⚠ تعذر الاتصال بالخادم. تحقق من اتصال الإنترنت.", duration=5000)
        except requests.exceptions.Timeout:
            self.notification_bar.show_message("⚠ انتهت مهلة الاتصال بالخادم. حاول لاحقًا.", duration=5000)
        except requests.exceptions.HTTPError as e:
            self.notification_bar.show_message(f"⚠ خطأ في الخادم: {e.response.status_code}", duration=5000)
        except requests.exceptions.RequestException:
            self.notification_bar.show_message("⚠ حدث خطأ أثناء التحقق من التحديث.", duration=5000)
        except ValueError:
            self.notification_bar.show_message("⚠ ملف التحديث غير صالح أو لا يمكن تحليله.", duration=5000)
        except Exception as e:
            self.notification_bar.show_message(f"⚠ خطأ غير متوقع: {str(e)}", duration=5000)
        finally:
            self._end_modal_dialog()

    def download_update(self, installer_url):
        import requests
        import os
        from PyQt5.QtWidgets import QMessageBox
        from PyQt5.QtCore import QStandardPaths
        from PyQt5.QtGui import QDesktopServices
        from PyQt5.QtCore import QUrl
        import socket

        def have_internet_connection():
            try:
                socket.create_connection(("8.8.8.8", 53), timeout=3)
                return True
            except OSError:
                return False

        if not have_internet_connection():
            self.notification_bar.show_message("⚠ لا يوجد اتصال بالإنترنت. لا يمكن تحميل التحديث.", duration=5000)
            return

        try:
            self.notification_bar.show_message("📦 جاري تحميل التحديث...", duration=4000)

            download_dir = QStandardPaths.writableLocation(QStandardPaths.TempLocation)
            installer_path = os.path.join(download_dir, "StMaryUpdater.exe")

            with requests.get(installer_url, stream=True, timeout=20) as r:
                r.raise_for_status()
                with open(installer_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)

            QMessageBox.information(
                self,
                "تم تحميل التحديث",
                "✅ تم تحميل التحديث بنجاح. سيتم الآن تثبيته، وسيتم إغلاق البرنامج.",
                QMessageBox.Ok
            )

            QDesktopServices.openUrl(QUrl.fromLocalFile(installer_path))
            QApplication.quit()

        except requests.exceptions.ConnectionError:
            self.notification_bar.show_message("⚠ تعذر الاتصال بالخادم. تحقق من اتصال الإنترنت.", duration=5000)
        except requests.exceptions.Timeout:
            self.notification_bar.show_message("⚠ انتهت مهلة التحميل. حاول مرة أخرى.", duration=5000)
        except requests.exceptions.HTTPError as e:
            self.notification_bar.show_message(f"⚠ خطأ في التحميل: {e.response.status_code}", duration=5000)
        except requests.exceptions.RequestException:
            self.notification_bar.show_message("⚠ حدث خطأ أثناء تحميل التحديث.", duration=5000)
        except Exception as e:
            self.notification_bar.show_message(f"⚠ خطأ غير متوقع: {str(e)}", duration=5000)

    def restart_app(self):
        """Restarts the current Python script with proper cleanup and hardcoded path."""
        import sys
        import subprocess
        
        # Show notification that app is restarting
        self.notification_bar.show_message("جاري إعادة تشغيل البرنامج...")
        
        # Process events to ensure the message is displayed
        QApplication.processEvents()
        
        try:
            # Close any open PowerPoint instances
            if self.is_powerpoint_open():
                try:
                    close_all_presentations_safe()
                except Exception:
                    pass
            
            # Delay slightly to allow resources to be released
            from time import sleep
            sleep(0.5)
            
            # Hardcoded file path for development
            script_path = r"F:\5dmt Shashat\Codes and Files\stmarymaadiliturgies.py"
            
            # Close the current application
            QApplication.quit()
            
            # Use subprocess to start a new instance
            subprocess.Popen([sys.executable, script_path])
            
            # Exit the current process
            sys.exit(0)
        except Exception as e:
            # If restart fails, show error and continue
            self.notification_bar.show_message(f"فشل إعادة التشغيل: {str(e)}")

    def setup_powerpoint_event_listener(self):
        """Set up a lightweight PowerPoint polling timer to clear glow after close."""
        try:
            self.last_open_presentations = set(get_open_presentations())
        except Exception:
            self.last_open_presentations = set()

        self._enable_ppt_polling = True
        self.ppt_check_timer = QTimer(self)
        self.ppt_check_timer.timeout.connect(self.check_powerpoint_changes)
        self.ppt_check_timer.start(2000)

    def check_powerpoint_changes(self):
        """Check if any PowerPoint presentations have been closed and update UI immediately"""
        if not getattr(self, "_enable_ppt_polling", False):
            return
        if not self.isActiveWindow():
            return
        try:
            # Get current presentations
            current_presentations = set(get_open_presentations())
            
            # If different from last check, refresh the buttons
            if current_presentations != self.last_open_presentations:
                # Check if رفع بخور عشية و باكر.pptx was closed
                import os
                baker_path = os.path.abspath(relative_path(r"رفع بخور عشية و باكر.pptx")).lower()
                was_open = any(path.lower() == baker_path for path in self.last_open_presentations)
                is_now_open = any(path.lower() == baker_path for path in current_presentations)
                
                if was_open and not is_now_open:
                    # Reset active source when the file is closed
                    self.active_presentation_source = None
                
                # Force a refresh without waiting
                self.refresh_button_states(skip_timer=True)
                self.last_open_presentations = current_presentations
        except Exception as e:
            print(f"Error checking PowerPoint changes: {e}")    

    def refresh_button_states(self, skip_timer=False):
        """
        Updates glow effects on buttons based on currently open presentations.
        Also closes any open SectionSelectionDialog if present.
        """
        import os

        frame2 = getattr(self, "frame2", None)
        if frame2 is None or self._rebuilding_main_frame:
            return

        # Get list of open presentations
        try:
            open_presentations = get_open_presentations()
        except Exception as e:
            print(f"Error reading open presentations: {e}")
            open_presentations = []
        
        # Map of buttons and their corresponding presentation files
        button_map = {
            "القداس": r"قداس.pptx",
            "قداس الطفل": r"قداس الطفل.pptx",
            "باكر": r"رفع بخور عشية و باكر.pptx",
            "عشية": r"رفع بخور عشية و باكر.pptx",
            "الإبصلمودية": r"الإبصلمودية.pptx",
            "المدائح": r"كتاب المدائح.pptx"
        }
        
        # Check if رفع بخور عشية و باكر.pptx is open
        baker_open = False
        baker_path = os.path.abspath(relative_path(r"رفع بخور عشية و باكر.pptx")).lower()
        if any(open_pres.lower() == baker_path for open_pres in open_presentations):
            baker_open = True
        
        # Find all buttons in the frame2 container
        try:
            for child in frame2.children():
                if isinstance(child, QFrame):
                    for btn_child in child.children():
                        if isinstance(btn_child, QLabel) and btn_child.text() in button_map:
                            button_text = btn_child.text()
                            container = btn_child.parent()
                            
                            # Use ABSOLUTE path for comparison
                            full_path = os.path.abspath(relative_path(button_map[button_text])).lower()
                            
                            # Special handling for باكر and عشية
                            if button_text in ["باكر", "عشية"]:
                                # Only add glow to the active button if the file is open
                                if baker_open and self.active_presentation_source == button_text:
                                    glow = QGraphicsDropShadowEffect(container)
                                    glow.setOffset(0)
                                    glow.setBlurRadius(30)
                                    glow.setColor(QColor(0, 255, 0))
                                    container.setGraphicsEffect(glow)
                                else:
                                    container.setGraphicsEffect(None)
                            else:
                                # Standard handling for other buttons
                                is_open = any(open_pres.lower() == full_path for open_pres in open_presentations)
                                if is_open:
                                    glow = QGraphicsDropShadowEffect(container)
                                    glow.setOffset(0)
                                    glow.setBlurRadius(30)
                                    glow.setColor(QColor(0, 255, 0))
                                    container.setGraphicsEffect(glow)
                                else:
                                    container.setGraphicsEffect(None)
        except RuntimeError:
            return

        if not skip_timer and getattr(self, "_enable_ppt_polling", False):
            if hasattr(self, "refresh_timer") and self.refresh_timer is not None:
                self.refresh_timer.start(1000)

    def focusInEvent(self, event):
        super().focusInEvent(event)
        self.refresh_button_states(skip_timer=True)

    def closeEvent(self, event):
        if hasattr(self, "refresh_timer") and self.refresh_timer is not None:
            self.refresh_timer.stop()
        if hasattr(self, "ppt_check_timer") and self.ppt_check_timer is not None:
            self.ppt_check_timer.stop()
        self._end_modal_dialog()
        super().closeEvent(event)

if __name__ == "__main__":
    app = QApplication(argv)

    # Show splash screen
    splash = ModernSplashScreen()
    splash.show()
    
    # Process events to make sure splash screen is displayed
    app.processEvents()
    
    # Create a QTimer to delay the main window
    from PyQt5.QtCore import QTimer
    timer = QTimer()
    timer.setSingleShot(True)
    
    def show_main_window():
        # Create and show main window
        window = MainWindow()
        window.show()
        
        # Close splash screen
        splash.close()
    
    # Set timer to trigger after 3000ms (3 seconds)
    timer.timeout.connect(show_main_window)
    timer.start(3000)

    exit(app.exec_())