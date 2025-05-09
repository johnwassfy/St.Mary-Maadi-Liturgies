import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def relative_path(relative_path):
    script_directory = os.path.dirname(os.path.abspath(__file__))
    absolute_path = os.path.join(script_directory, relative_path)
    return absolute_path

def extract_section_info(file_sheet_pairs, excel_file):
    try:
        # Load the existing Excel file
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = openpyxl.Workbook()

    for pptx_file, sheet_name in file_sheet_pairs:
        # Extract XML from PowerPoint file
        with zipfile.ZipFile(pptx_file, 'r') as zip_ref:
            zip_ref.extract('ppt/presentation.xml', 'temp_extract')
        
        # Parse XML
        tree = ET.parse('temp_extract/ppt/presentation.xml')
        root = tree.getroot()
        
        # Namespace for PowerPoint XML
        ns = {'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
              'p14': 'http://schemas.microsoft.com/office/powerpoint/2010/main'}

        # List to hold section info for the current presentation
        section_info = []
        
        # Iterate through sections to extract section info
        slide_count = 0
        for section in root.findall('.//p14:section', ns):
            section_name = section.get('name')
            slide_ids = section.findall('.//p14:sldId', ns)
            if slide_ids:
                first_slide_id = int(slide_ids[0].get('id'))
                last_slide_id = int(slide_ids[-1].get('id'))
                section_info.append({'Section Name': section_name,
                                     'First Slide Number': slide_count + 1,
                                     'Last Slide Number': slide_count + len(slide_ids),
                                     'Number of Slides': len(slide_ids)})
                slide_count += len(slide_ids)
        
        # Convert section_info to DataFrame
        df = pd.DataFrame(section_info)

        # Get or create the worksheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Clear existing contents of the sheet
            ws.delete_rows(1, ws.max_row)
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        # Write DataFrame to Excel sheet without index
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        ws.sheet_view.rightToLeft = True
        
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1).value = "Design"
        ws.cell(row=last_row, column=2).value = True

        # Save the Excel file
        wb.save(excel_file)

def extract_section_info2(file_sheet_pairs, excel_file, progress_callback=None):
    try:
        # Load the existing Excel file
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = openpyxl.Workbook()

    total_files = len(file_sheet_pairs)  # Total number of files to process

    for index, (pptx_file, sheet_name) in enumerate(file_sheet_pairs):
        # Update progress after starting each file
        if progress_callback:
            progress = int((index / total_files) * 50)  # Scale to 50% for this step
            progress_callback(progress)

        # Extract XML from PowerPoint file
        with zipfile.ZipFile(pptx_file, 'r') as zip_ref:
            zip_ref.extract('ppt/presentation.xml', 'temp_extract')

        # Parse XML
        tree = ET.parse('temp_extract/ppt/presentation.xml')
        root = tree.getroot()

        # Namespace for PowerPoint XML
        ns = {'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
              'p14': 'http://schemas.microsoft.com/office/powerpoint/2010/main'}

        # List to hold section info for the current presentation
        section_info = []

        # Iterate through sections to extract section info
        slide_count = 0
        for section in root.findall('.//p14:section', ns):
            section_name = section.get('name')
            section_id = section.get('id')  # Use the existing 'id' attribute
            slide_ids = section.findall('.//p14:sldId', ns)
            if slide_ids:
                first_slide_id = int(slide_ids[0].get('id'))
                last_slide_id = int(slide_ids[-1].get('id'))
                section_info.append({
                    'Section Name': section_name,
                    'Section ID': section_id,
                    'First Slide Number': slide_count + 1,
                    'Last Slide Number': slide_count + len(slide_ids),
                    'Number of Slides': len(slide_ids)
                })
                slide_count += len(slide_ids)

        # Convert section_info to DataFrame
        df = pd.DataFrame(section_info)

        # Get or create the worksheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)  # Clear existing contents
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        # Write DataFrame to Excel sheet without index
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        ws.sheet_view.rightToLeft = True

        # Save the Excel file
        wb.save(excel_file)

    # Ensure progress reaches 50% after extracting sections
    if progress_callback:
        progress_callback(50)
